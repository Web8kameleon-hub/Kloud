from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill

from lagter_v1_models import LagterPayload
from lagter_v1_sources import collect_real_snapshot


HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBHEADER_FONT = Font(color="1F4E78", bold=True)


class LagterV1ExcelBuilder:
    def __init__(self) -> None:
        self.generated_at = datetime.now(timezone.utc).isoformat()

    @staticmethod
    def _clamp01(value: float) -> float:
        return max(0.0, min(1.0, float(value)))

    def build_payload(self) -> Dict[str, Any]:
        snapshot = collect_real_snapshot()

        service_coverage = self._clamp01(snapshot.services_total / 75.0) if snapshot.services_total else 0.0
        lab_coverage = self._clamp01(snapshot.labs_total / 23.0) if snapshot.labs_total else 0.0
        quality_proxy = self._clamp01(1.0 - (snapshot.memory_percent / 100.0))

        error_rate = (snapshot.telemetry_errors / snapshot.telemetry_events) if snapshot.telemetry_events else 0.0
        warning_rate = (snapshot.telemetry_warnings / snapshot.telemetry_events) if snapshot.telemetry_events else 0.0
        stability_proxy = self._clamp01(1.0 - ((snapshot.cpu_percent + snapshot.memory_percent) / 200.0))

        law_i_rate = self._clamp01((service_coverage + lab_coverage) / 2.0)
        law_ii_rate = self._clamp01(stability_proxy)
        law_iii_rate = self._clamp01(1.0 - error_rate)

        if snapshot.system_metric_points:
            points = snapshot.system_metric_points[-5:]
        else:
            points = [
                {
                    "timestamp": snapshot.generated_at,
                    "cpu_percent": snapshot.cpu_percent,
                    "ram_percent": snapshot.memory_percent,
                    "disk_percent": snapshot.disk_percent,
                }
            ]

        sketch_points: List[Dict[str, Any]] = []
        for idx, point in enumerate(points, start=1):
            bio = self._clamp01(1.0 - (float(point.get("ram_percent", 0.0)) / 100.0))
            behavior = self._clamp01(1.0 - (float(point.get("cpu_percent", 0.0)) / 100.0))
            ambient = self._clamp01(1.0 - (float(point.get("disk_percent", 0.0)) / 100.0))
            tension = self._clamp01((float(point.get("cpu_percent", 0.0)) + float(point.get("ram_percent", 0.0))) / 200.0)
            sketch_points.append(
                {
                    "day": f"P{idx}",
                    "bio": bio,
                    "behavior": behavior,
                    "ambient": ambient,
                    "tension": tension,
                }
            )

        enigma_registry: List[Dict[str, Any]] = []
        if snapshot.telemetry_errors > 0:
            enigma_registry.append(
                {
                    "enigma_id": "ENI-REAL-ERR",
                    "title": "Error events detected in telemetry",
                    "status": "open",
                    "confidence": self._clamp01(1.0 - error_rate),
                    "hypothesis": f"{snapshot.telemetry_errors} error events found in telemetry logs; investigate service instability.",
                }
            )
        if snapshot.telemetry_warnings > 0:
            enigma_registry.append(
                {
                    "enigma_id": "ENI-REAL-WARN",
                    "title": "Warning events detected in telemetry",
                    "status": "testing",
                    "confidence": self._clamp01(1.0 - warning_rate),
                    "hypothesis": f"{snapshot.telemetry_warnings} warning events suggest transient tension in pipeline.",
                }
            )

        return {
            "generated_at": snapshot.generated_at,
            "version": "v1",
            "kpis": [
                {"name": "service_coverage", "target": 1.0, "actual": round(service_coverage, 4)},
                {"name": "lab_coverage", "target": 1.0, "actual": round(lab_coverage, 4)},
                {"name": "telemetry_quality", "target": 0.98, "actual": round(self._clamp01(1.0 - error_rate), 4)},
                {"name": "stability_proxy", "target": 0.75, "actual": round(stability_proxy, 4)},
                {"name": "quality_proxy", "target": 0.8, "actual": round(quality_proxy, 4)},
            ],
            "law_checks": [
                {"law": "Ligji I", "description": "Dy pole + zinxhir", "pass_rate": round(law_i_rate, 4)},
                {"law": "Ligji II", "description": "Balancë para vendimit", "pass_rate": round(law_ii_rate, 4)},
                {"law": "Ligji III", "description": "Enigma si sinjal", "pass_rate": round(law_iii_rate, 4)},
            ],
            "enigma_registry": enigma_registry,
            "sketch_points": sketch_points,
        }

    def validate_payload(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        validated = LagterPayload.model_validate(payload)
        return validated.model_dump()

    def process_map(self) -> List[Dict[str, Any]]:
        return [
            {
                "step": "capture_signals",
                "input": "bio + behavior + ambient",
                "output": "aligned_observations",
                "controls": ["quality_score", "time_sync"],
            },
            {
                "step": "build_dual_signals",
                "input": "aligned_observations",
                "output": "dual_signal_frames",
                "controls": ["law_i_dual_pole", "chain_coverage"],
            },
            {
                "step": "compute_tension",
                "input": "dual_signal_frames",
                "output": "tension_index",
                "controls": ["stability_window", "drift_check"],
            },
            {
                "step": "balance_then_decide",
                "input": "tension_index",
                "output": "action_state",
                "controls": ["law_ii_balance", "persistent_vs_transient"],
            },
            {
                "step": "enigma_registry",
                "input": "low_confidence_decisions",
                "output": "enigma_cases",
                "controls": ["law_iii_enigma_signal", "hypothesis_tracking"],
            },
            {
                "step": "excel_reporting",
                "input": "action_state + enigma_cases",
                "output": "lagter_v1_dashboard.xlsx",
                "controls": ["audit_trail", "version_tag"],
            },
        ]

    def build_workbook(self, payload: Dict[str, Any] | None = None) -> Workbook:
        data = self.validate_payload(payload or self.build_payload())
        wb = Workbook()

        ws_overview = wb.active
        ws_overview.title = "Overview"
        self._build_overview(ws_overview, data)

        ws_kpi = wb.create_sheet("KPI")
        self._build_kpi(ws_kpi, data["kpis"])

        ws_laws = wb.create_sheet("LawCompliance")
        self._build_laws(ws_laws, data["law_checks"])

        ws_enigma = wb.create_sheet("EnigmaRegistry")
        self._build_enigma(ws_enigma, data["enigma_registry"])

        ws_sketch = wb.create_sheet("Sketches")
        self._build_sketches(ws_sketch, data["sketch_points"])

        return wb

    def write_workbook(self, output_path: str | Path, payload: Dict[str, Any] | None = None) -> Path:
        destination = Path(output_path)
        destination.parent.mkdir(parents=True, exist_ok=True)
        workbook = self.build_workbook(payload)
        workbook.save(destination)
        return destination

    def _style_header(self, ws, row: int, headers: List[str]) -> None:
        for col_index, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_index, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _build_overview(self, ws, data: Dict[str, Any]) -> None:
        ws["A1"] = "L.A.G.T.E.R v1 - Excel Core"
        ws["A1"].font = Font(size=16, bold=True, color="1F4E78")

        lines = [
            ("Generated At", data["generated_at"]),
            ("Version", data["version"]),
            ("Module", "LAGTER v1 integrated with Excel Core"),
            ("Scope", "Biology - Behavior - Ambient chain"),
            ("Law Gate", "Ligji I / Ligji II / Ligji III"),
        ]

        row = 3
        for label, value in lines:
            ws.cell(row=row, column=1, value=label).font = SUBHEADER_FONT
            ws.cell(row=row, column=2, value=value)
            row += 1

        ws["A10"] = "Flow Sketch"
        ws["A10"].font = SUBHEADER_FONT
        ws["A11"] = "Signal -> Counter-Signal -> Tension -> Balance -> Enigma -> Audited Decision"
        ws["A11"].alignment = Alignment(wrap_text=True)

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 52

    def _build_kpi(self, ws, kpis: List[Dict[str, Any]]) -> None:
        headers = ["KPI", "Target", "Actual", "Delta"]
        self._style_header(ws, 1, headers)

        row = 2
        for item in kpis:
            delta = round(item["actual"] - item["target"], 4)
            ws.cell(row=row, column=1, value=item["name"])
            ws.cell(row=row, column=2, value=item["target"])
            ws.cell(row=row, column=3, value=item["actual"])
            ws.cell(row=row, column=4, value=delta)
            row += 1

        chart = BarChart()
        chart.title = "Target vs Actual"
        chart.y_axis.title = "Score"
        chart.x_axis.title = "KPI"

        values = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=1 + len(kpis))
        categories = Reference(ws, min_col=1, min_row=2, max_row=1 + len(kpis))
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 8
        chart.width = 18
        ws.add_chart(chart, "F2")

        for col in ["A", "B", "C", "D"]:
            ws.column_dimensions[col].width = 24 if col == "A" else 14

    def _build_laws(self, ws, laws: List[Dict[str, Any]]) -> None:
        headers = ["Law", "Description", "Pass Rate"]
        self._style_header(ws, 1, headers)

        for idx, law in enumerate(laws, start=2):
            ws.cell(row=idx, column=1, value=law["law"])
            ws.cell(row=idx, column=2, value=law["description"])
            ws.cell(row=idx, column=3, value=law["pass_rate"])

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 48
        ws.column_dimensions["C"].width = 14

    def _build_enigma(self, ws, records: List[Dict[str, Any]]) -> None:
        headers = ["Enigma ID", "Title", "Status", "Confidence", "Hypothesis"]
        self._style_header(ws, 1, headers)

        for idx, record in enumerate(records, start=2):
            ws.cell(row=idx, column=1, value=record["enigma_id"])
            ws.cell(row=idx, column=2, value=record["title"])
            ws.cell(row=idx, column=3, value=record["status"])
            ws.cell(row=idx, column=4, value=record["confidence"])
            ws.cell(row=idx, column=5, value=record["hypothesis"])

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 42
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 56

    def _build_sketches(self, ws, points: List[Dict[str, Any]]) -> None:
        headers = ["Day", "Bio", "Behavior", "Ambient", "Tension"]
        self._style_header(ws, 1, headers)

        for row_index, point in enumerate(points, start=2):
            ws.cell(row=row_index, column=1, value=point["day"])
            ws.cell(row=row_index, column=2, value=point["bio"])
            ws.cell(row=row_index, column=3, value=point["behavior"])
            ws.cell(row=row_index, column=4, value=point["ambient"])
            ws.cell(row=row_index, column=5, value=point["tension"])

        line_chart = LineChart()
        line_chart.title = "Chain Sketch (Bio/Behavior/Ambient/Tension)"
        line_chart.y_axis.title = "Index"
        line_chart.x_axis.title = "Day"

        data = Reference(ws, min_col=2, min_row=1, max_col=5, max_row=1 + len(points))
        categories = Reference(ws, min_col=1, min_row=2, max_row=1 + len(points))
        line_chart.add_data(data, titles_from_data=True)
        line_chart.set_categories(categories)
        line_chart.height = 10
        line_chart.width = 20
        ws.add_chart(line_chart, "G2")

        for col in ["A", "B", "C", "D", "E"]:
            ws.column_dimensions[col].width = 14


def build_lagter_v1_excel(output_path: str | Path, payload: Dict[str, Any] | None = None) -> Path:
    builder = LagterV1ExcelBuilder()
    return builder.write_workbook(output_path=output_path, payload=payload)
